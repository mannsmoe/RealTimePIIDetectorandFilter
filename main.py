import xlrd
import xlwt 
import openpyxl
import os

print("============================")
print("Loading and reading SIMPLE excel file data")
print("============================")
print("Loading the Excel File")
excel_file=xlrd.open_workbook("mk.xlsx")
print("This Excel File has: ",excel_file.nsheets," Sheets in it")
print("sheets names are: ",excel_file.sheet_names())
sheet1=excel_file.sheet_by_index(0)
print(sheet1.row_values(0))
cell31=sheet1.cell(4,1)
print(cell31)
print(" \n")
print("Creating am excel file for the output")
ouptput=xlwt.Workbook(encoding="utf-8")
sheet1=ouptput.add_sheet("First_Sheet_in_the_Excel_File")
sheet2=ouptput.add_sheet("Second_Sheet_in_the_Excel_File")
sheet3=ouptput.add_sheet("Third_Sheet_in_the_Excel_File")
sheet1.write(0,0, "cell31")
sheet2.write(0,0, "cell31")
sheet3.write(0,0, "cell31")
ouptput.save("output_file.xls")
print("Output excel file created")

print('END')
print("============================")
print("============================")

















print(" \n" *5)
print("============================")
print("Detecting any email address") 
print("============================")
for line in open("mk.txt"):
    if "@" and ".com" or ".org" in line:
        found_email=line
        print(found_email)
        
        
        
print('END')
print("============================")
print("============================")








print(" \n" *5)
print("============================")
print("Loading and reading LARGE excel file data")
print("============================")
data = openpyxl.load_workbook('mk2.xlsx')
x1=type(data)
print(x1)
x2=data.get_sheet_names()
print(x2)
sheet_1=data.get_sheet_by_name('new sheet name')
print(sheet_1)
x3=sheet_1['A1'].value
print(x3)
x4=sheet_1['D4'].value
print(x4)
sheet_1['A3'].value='Steve'
x5=data.save('mk2.xlsx')
print(x5)
x6=sheet_1.title
print(x6)
sheet_1.title='new sheet name'
data.save('mk2.xlsx')
x7=sheet_1.cell(row=3, column=1).value
print(x7)
for i in range(1, 6): print(sheet_1.cell(row=i, column=2).value)
print(sheet_1.max_row)
print(sheet_1.max_column)


try: 
    from openpyxl.cell import get_column_letter
except ImportError:
    from openpyxl.utils import get_column_letter
    
    
print(openpyxl.utils.get_column_letter(2))
print(openpyxl.utils.column_index_from_string('D'))
data.create_sheet(title='added_sheet', index=1)
data.save('mk3.xlsx')
sheet_1.row_dimensions[1].height=200
sheet_1.column_dimensions['A'].width=70
data.save('mk4.xlsx')
from openpyxl.styles import Font
sheet_1['C1'].font=Font(sz=32, bold=True, italic=True)
data.save('mk5.xlsx')

data=openpyxl.Workbook()
s1=data.create_sheet('my_s')
import random
for i in range(1, 110): 
    s1['A'+str(i)].value=random.randint(1,100)
    s1['B'+str(i)].value=random.randint(1,100)
    s1['C'+str(i)].value=random.randint(1,100)
data.save('mk6.xlsx')

ws = data.active
for i in range(10):
    ws.append([i])
    
from openpyxl.chart import BarChart, Reference, Series
values = Reference(ws, min_col=1, min_row=1, max_col=1, max_row=10)
chart = BarChart()
chart.add_data(values)
ws.add_chart(chart, "E15")
data.save("SampleChart.xlsx")
#openpyxl.chart.reference(s1, (1,1), (10,1))








print(" \n" *5)
print("============================")
print("Searching a LARGE excel file for PII")
print("============================")
# Loading the raw data file
d1 = openpyxl.load_workbook('pii1.xlsx')
print(type(d1))
#scanning the excel file contents
#print(d1.get_sheet_names())
shsarray=d1.get_sheet_names()
shn=0;
for sheet in d1.worksheets:
    shn = shn + 1
print("This excel file contains " + str(shn) + " sheets")
#print(shsarray)
for i in range(0, len(shsarray)): print('Sheet #' + str(i+1) + ' is: ' + str(shsarray[i]))

for i in range(0, len(shsarray)): print(shsarray[i])

for i in range (0, shn):                              #Loop for the sheets
    print('For the sheet #' + str(i+1) + ':')
    sh_data=d1.get_sheet_by_name(shsarray[i])
    rn=sh_data.max_row
    cn=sh_data.max_column
    flag1=1
    print('Number of rows is:' + str(rn))             
    print('Number of rows is:' + str(cn))
    for j in range (1, cn+1):                         #Loop for the columns
        x7=sh_data.cell(row=1, column=j).value
        #if x7.lower() == 'email':                      #Optimizing the search time
        if x7 != '':
            if x7 == 'email' or x7 == 'Email' or x7 == 'EMAIL':
                flag1=0
                bad_column=j
                print('Email column detected, located in the column #' + str(bad_column) + ' ==> ' + openpyxl.utils.get_column_letter(bad_column))
                for k in range(2, rn+1):                  #Loop for cleaning
                    sh_data.cell(row=k, column=bad_column).value=''
                    d1.save('pii1_email_clear.xlsx')
                    print('Cleaning the emails is done and saved !')
print(x7)


if flag1 == 1:
    print('No email title detected in this sheet ...')
    print('Starting deep searching for email PII depending on the pattern...')
    import pyperclip, re
    # email regex:
    emailRegex = re.compile(r'''(
        [a-zA-Z0-9._%+-]+ # username
        @ # @ symbol
        [a-zA-Z0-9.-]+ # domain name
        (\.[a-zA-Z]{2,4}) # dot-something
        )''', re.VERBOSE)
    matches = []
    suspecious_col=[]
    print ('\n'*5)
    d1 = openpyxl.load_workbook('pii1.xlsx')
    print(type(d1))
    #scanning the excel file contents
    shsarray=d1.get_sheet_names()
    shn=0;
    for sheet in d1.worksheets:
        shn = shn + 1
    print("This excel file contains " + str(shn) + " sheets")
    for i in range(0, len(shsarray)): print('Sheet #' + str(i+1) + ' is: ' + str(shsarray[i]))
    for i in range(0, len(shsarray)): print(shsarray[i])
    for i in range (0, shn):                              #Loop for the sheets
        print('For the sheet #' + str(i+1) + ':')
        sh_data=d1.get_sheet_by_name(shsarray[i])
        rn=sh_data.max_row
        cn=sh_data.max_column
        print('Number of rows is:' + str(rn))             
        print('Number of rows is:' + str(cn))
        noTestRecords = 10
        threshPIISus = 0.7
        noThreshPIISus = noTestRecords * threshPIISus
        for j in range (1, cn+1):                         #Loop for the columns
            for jj in range (1, 4):                       #Loop for the first four rows 
                x7=sh_data.cell(row=jj, column=j).value
                for groups in emailRegex.findall(str(x7)):         #The condition
                    matches.append(groups[0])
                    
                if len(matches) > 0:
                    bad_column=j
                    suspecious_col.append(bad_column)
                    #suspecious_col[j]=bad_column
                    bad_row=jj                              #is there a possibility for horizental data pattern? 
                    print('Email pattern found in column #' + str(bad_column) + ' ==> ' + openpyxl.utils.get_column_letter(bad_column) + ', and row # ' + str(bad_row))
                
            #testing if within the first four rows there is a suspicious column
            if len(suspecious_col) >= noThreshPIISus:  #75% case
                print(' More than len(suspecious_col) the cells of the column #' + str(bad_column) + 'contain email addresses, this column must be filtered ...' )
                print('Cleaning in progress ... ')
                for k in range(1, rn+1):                  #Loop for cleaning
                    sh_data.cell(row=k, column=bad_column).value=''
                    d1.save('pii1_email_clear.xlsx')
                print('Cleaning the emails is done and saved !')
                    
                    
            if len(suspecious_col) < 3: #75% case
                print(' 50% or less of the cells of the column #' + str(bad_column) + 'contain email addresses, operator must decide ...' )
                print('please hit c for cleaning, d for start deep search, or s to skip')
                controlv = input().lower()




    
                
                






